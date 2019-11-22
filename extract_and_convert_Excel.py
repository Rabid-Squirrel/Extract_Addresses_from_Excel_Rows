#Project Working with Excel
from openpyxl import load_workbook
from openpyxl import Workbook
#import xlwt
def open_data():
    wb=load_workbook('Sports Fields MH.xlsx')
    result=[]
    ws=wb.worksheets[0]
    index=0
    for row in ws.iter_rows() :
        result.append([cell.value for cell in row])
        #print(index)
        index +=1
        if index==115:
            break
    print(index)
    print(result)
    #print(#len(result))
    row_need=[]
    for index,row in enumerate(result):
        print(index,row)
        if index > 0 :#first row not write
            try:
                for r in row[0].split('\n'):
                    row_need.append(r)
            except:
                print("Error 0")
                continue
            try:
                for r in row[1].split('\n'):
                    row_need.append(r)
            except:
                print("Error 1")
                continue
            try:
                for r in row[2].split('\n'):
                    row_need.append(r)
            except:
                print("Error 2")
                continue
            try:
                for r in row[3].split('\n'):
                    row_need.append(r)
            except:
                print("Error 3")
                continue
    row_need_finish=[]
    row_need_finish.append(['street_address','city','state','zip_code'])

    print('=====================================================')
    for i,row in enumerate(row_need):
        try:
            row_temp=row.strip().split(',')
            print("row_temp: ",row_temp)
            state=None
            city =  row_temp[-2].strip()
            zip_code = row_temp[-1].strip()
            street_address=','.join(row_temp[0:-1])
            print('street_address:{},city:{},state:{},zip_code:{}'.format(street_address,city,state,zip_code))
        except:
            print('ERROR: ',row_temp)
            continue

        row_need_finish.append([street_address,city,state,zip_code])
    print(row_need_finish)
    return row_need_finish

def write_data_excel(row_need_finish):
    wb = Workbook()
    ws = wb.active
    for row in row_need_finish:
        ws.append(row)
    wb.save("Sports Fields MH address extracted.xlsx")

def main():
    row_need_finish = open_data()
    write_data_excel(row_need_finish)

if __name__ == '__main__':
    main()